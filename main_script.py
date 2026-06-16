"""Append live market and FII/DII data to Forex_Insights.xlsx."""

import json
import re
import threading
import time
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any
from urllib.parse import quote

import websocket
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from notifier import send_slack_summary
from drive_uploader import upload_to_drive

EXCEL_FILE = "Forex_Insights.xlsx"
US_SHEET = "US_Indices"
LAUNCH_DELAY = 0.6
MAX_RETRIES = 3
RETRY_DELAY = 5.0
FPI_DEBT_SUBTOTALS = 3
TV_CANDLE_COUNT = 5

SYMBOL_MARKET = {
    "DXY": "FX", "EURUSD": "FX", "USDJPY": "FX", "GBPUSD": "FX",
    "USDCAD": "FX", "USDSEK": "FX", "USDCHF": "FX", "USDCNH": "FX",
    "USDMXN": "FX", "USDKRW": "FX", "USDBRL": "FX", "USDINR": "FX",
    "USDNPR": "FX", "NPRUSD": "FX",
    "OIL": "OIL",
    "XAUUSD": "GOLD", "XAUFUT": "GOLD", "XAGUSD": "GOLD",
    "NIFTY": "INDIA", "SENSEX": "INDIA",
    "NDX": "US", "SPX": "US",
}

SESSION_INFO_NPT = {
    "FX":    {
        "open_summer":  "02:45",
        "close_summer": "02:44+1",
        "open_winter":  "03:45",
        "close_winter": "03:44+1",
        "note": "NY close: 17:00 EDT=02:45 NPT (summer)  /  17:00 EST=03:45 NPT (winter)",
    },
    "OIL":   {
        "open_summer": "02:45", "close_summer": "02:44+1",
        "open_winter": "03:45", "close_winter": "03:44+1",
        "note": "ICE Brent follows NY close same as FX",
    },
    "GOLD":  {
        "open_summer": "02:45", "close_summer": "02:44+1",
        "open_winter": "03:45", "close_winter": "03:44+1",
        "note": "COMEX NY close",
    },
    "INDIA": {
        "open_summer":  "09:30",
        "close_summer": "15:45",
        "open_winter":  "09:30",
        "close_winter": "15:45",
        "note": "NSE 09:15-15:30 IST  =  09:30-15:45 NPT  (IST + 15 min = NPT)",
    },
    "US":    {
        "open_summer":  "19:15",
        "close_summer": "01:45+1",
        "open_winter":  "20:15",
        "close_winter": "02:45+1",
        "note": "NYSE 09:30-16:00 EDT=19:15-01:45 NPT (summer)  /  EST=20:15-02:45 NPT (winter)",
    },
}

INDIA_TZ_HOURS: float = 5.5


def _dst_active() -> bool:
    """Return True during approximate US DST period (March-October)."""
    return 3 <= datetime.utcnow().month <= 10


def session_close_utc_hour(market: str) -> int:
    """Return the UTC hour after which a market's daily bar is complete."""
    summer = _dst_active()
    if market in ("FX", "OIL", "GOLD"):
        return 21 if summer else 22
    if market == "INDIA":
        return 10
    if market == "US":
        return 20 if summer else 21
    return 21


def market_tz_offset(market: str) -> float:
    """Return the UTC offset (hours) needed to recover the correct calendar date
    from a TradingView daily candle timestamp.

    TradingView stamps each daily bar at its *open* time in the exchange's
    reference timezone:

    * FX / OIL / GOLD  — bar opens at NY close (21:00 UTC summer, 22:00 UTC
      winter).  Adding (24 - close_hour) h nudges the timestamp forward to the
      next UTC midnight, which is the conventional market-date label.  This is
      safe because any *completed* bar has ts < cutoff, so ts + offset ≤ 23:59
      UTC and never flips to the following day.
    * INDIA  — bar opens at NSE midnight IST (= UTC - 5.5 h on the previous
      UTC day); adding +5.5 h recovers the IST date.
    * US    — bar opens at NYSE midnight EDT/EST; subtracting the EDT/EST
      offset recovers the ET date.
    """
    if market in ("FX", "OIL", "GOLD"):
        return float(24 - session_close_utc_hour(market))
    if market == "INDIA":
        return INDIA_TZ_HOURS
    if market == "US":
        return -4.0 if _dst_active() else -5.0
    return 0.0


def candle_market_date(ts: float, key: str) -> str:
    """Return the exchange-local calendar date string for a candle timestamp."""
    market = SYMBOL_MARKET.get(key, "FX")
    offset = market_tz_offset(market)
    local_dt = datetime.utcfromtimestamp(ts) + timedelta(hours=offset)
    return local_dt.strftime("%Y-%m-%d")

SYMBOLS = {
    "DXY": "TVC:DXY",
    "EURUSD": "FX_IDC:EURUSD",
    "USDJPY": "FX_IDC:USDJPY",
    "GBPUSD": "FX_IDC:GBPUSD",
    "USDCAD": "FX_IDC:USDCAD",
    "USDSEK": "FX_IDC:USDSEK",
    "USDCHF": "FX_IDC:USDCHF",
    "USDCNH": "FX_IDC:USDCNH",
    "USDMXN": "FX_IDC:USDMXN",
    "USDKRW": "FX_IDC:USDKRW",
    "USDBRL": "FX_IDC:USDBRL",
    "USDINR": "FX_IDC:USDINR",
    "USDNPR": "FX_IDC:USDNPR",
    "NPRUSD": "FX_IDC:NPRUSD",
    "OIL": "UKOIL",
    "XAUUSD": "FX_IDC:XAUUSD",
    "XAUFUT": "COMEX:GC1!",
    "XAGUSD": "FX_IDC:XAGUSD",
    "NIFTY": "NSE:NIFTY",
    "SENSEX": "BSE:SENSEX",
    "NDX": "NASDAQ:NDX",
    "SPX": "SP:SPX",
}

SYMBOL_PAGE = {
    "TVC:DXY": "TVC-DXY",
    "UKOIL": "UKOIL",
    "COMEX:GC1!": "COMEX-GC1",
    "NSE:NIFTY": "NSE-NIFTY",
    "BSE:SENSEX": "BSE-SENSEX",
    "NASDAQ:NDX": "NASDAQ-NDX",
    "SP:SPX": "SP-SPX",
    "FX_IDC:EURUSD": "FX_IDC-EURUSD",
    "FX_IDC:USDJPY": "FX_IDC-USDJPY",
    "FX_IDC:GBPUSD": "FX_IDC-GBPUSD",
    "FX_IDC:USDCAD": "FX_IDC-USDCAD",
    "FX_IDC:USDSEK": "FX_IDC-USDSEK",
    "FX_IDC:USDCHF": "FX_IDC-USDCHF",
    "FX_IDC:USDCNH": "FX_IDC-USDCNH",
    "FX_IDC:USDMXN": "FX_IDC-USDMXN",
    "FX_IDC:USDKRW": "FX_IDC-USDKRW",
    "FX_IDC:USDBRL": "FX_IDC-USDBRL",
    "FX_IDC:USDINR": "FX_IDC-USDINR",
    "FX_IDC:USDNPR": "FX_IDC-USDNPR",
    "FX_IDC:NPRUSD": "FX_IDC-NPRUSD",
    "FX_IDC:XAUUSD": "FX_IDC-XAUUSD",
    "FX_IDC:XAGUSD": "FX_IDC-XAGUSD",
}

TV_HEADERS = {
    "Origin": "https://www.tradingview.com",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
}

MC_FII_WIDGET = "https://www.moneycontrol.com/mc/widget/fiidii"
NSE_FII_URL = "https://www.nseindia.com/api/fiidiiTradeReact"
NSDL_FPI_URL = "https://www.fpi.nsdl.co.in/web/Reports/Latest.aspx"
HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json,text/html,*/*",
}

DXY_PAIRS = [
    "DXY", "EURUSD", "USDJPY", "GBPUSD", "USDCAD",
    "USDSEK", "USDCHF", "USDCNH", "USDMXN", "USDKRW", "USDBRL",
]

FII_COLUMN_KEYS = (
    (13, "fii_net"),
    (14, "dii_net"),
    (15, "fii_fut_net"),
    (16, "fii_opt_net"),
    (17, "equity"),
    (18, "debt"),
)

UP_FILL = PatternFill("solid", start_color="E2EFDA")
DOWN_FILL = PatternFill("solid", start_color="FCE4D6")
TITLE_FILL = PatternFill("solid", start_color="356854")
TITLE_FONT = Font(bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
SUBHDR_FONT = Font(name="Arial", size=10, bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


@dataclass
class AppState:
    """Runtime caches for fetched market and FII data."""

    market: dict[str, dict[str, Any]] = field(default_factory=dict)
    fii: dict[str, Any] = field(default_factory=dict)
    run_at: datetime | None = None
    _lock: threading.Lock = field(default_factory=threading.Lock, repr=False)


state = AppState()


def tradingview_url(tv_symbol: str) -> str:
    """Build the TradingView WebSocket URL for a symbol."""
    page = SYMBOL_PAGE.get(tv_symbol, tv_symbol.replace(":", "-").replace("!", ""))
    today = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    return (
        f"wss://data.tradingview.com/socket.io/websocket"
        f"?from=symbols%2F{quote(page, safe='')}%2F"
        f"&date={quote(today, safe='')}&auth=sessionid"
    )


def send_socket_payload(ws: websocket.WebSocket, payload: dict) -> None:
    """Send a JSON payload using TradingView socket framing."""
    msg = json.dumps(payload)
    ws.send(f"~m~{len(msg)}~m~{msg}")


def parse_socket_message(message: str) -> list[Any]:
    """Split and parse a TradingView socket message into JSON objects."""
    parts = re.split(r"~m~\d+~m~", message)
    parsed = []
    for part in parts:
        if part.strip():
            try:
                parsed.append(json.loads(part))
            except json.JSONDecodeError:
                parsed.append(part)
    return parsed


def last_completed_candles(key: str, ordered: list[dict[str, Any]]) -> tuple[dict, dict]:
    """Return (latest_complete, previous) based on each market's session close time.

    Filters out any bar whose session has not yet closed at script run time.
    DST is accounted for: FX/OIL/GOLD/US close one UTC hour later in winter.
    Weekends are naturally absent from TradingView exchange-traded daily data.
    """
    market = SYMBOL_MARKET.get(key, "FX")
    close_hour = session_close_utc_hour(market)
    now_utc = datetime.utcnow()
    cutoff = now_utc.replace(hour=close_hour, minute=0, second=0, microsecond=0)
    if now_utc < cutoff:
        cutoff -= timedelta(days=1)
    complete = [c for c in ordered if datetime.utcfromtimestamp(c["ts"]) < cutoff]
    if not complete:
        complete = ordered
    latest = complete[-1]
    previous = complete[-2] if len(complete) > 1 else complete[-1]
    return latest, previous


def store_candles(key: str, candles: list[dict[str, Any]]) -> None:
    """Persist the last completed daily candle for a symbol into application state."""
    ordered = sorted(candles, key=lambda row: row["ts"])
    latest, previous = last_completed_candles(key, ordered)
    market_date = candle_market_date(latest["ts"], key)
    payload = {
        "date": market_date,
        "open": latest["open"],
        "high": latest["high"],
        "low": latest["low"],
        "close": latest["close"],
        "prev_close": previous["close"],
    }
    with state._lock:
        state.market[key] = payload
    print(f"  OK {key:10s}  {market_date}  C:{latest['close']}")


def fetch_tradingview_symbol(key: str, tv_symbol: str, attempt: int = 1) -> None:
    """Fetch daily candles for one symbol, with backoff on HTTP 429."""
    candles: list[dict[str, Any]] = []
    got_429 = threading.Event()
    done = threading.Event()
    session_id = f"cs_{abs(hash(tv_symbol + str(attempt))) % 99999:05d}"

    def on_open(ws: websocket.WebSocket) -> None:
        send_socket_payload(ws, {"m": "set_auth_token", "p": ["unauthorized_user_token"]})
        send_socket_payload(ws, {"m": "chart_create_session", "p": [session_id, ""]})
        send_socket_payload(
            ws,
            {
                "m": "resolve_symbol",
                "p": [
                    session_id,
                    "sds_sym_1",
                    f'={{"symbol":"{tv_symbol}","adjustment":"splits"}}',
                ],
            },
        )
        send_socket_payload(
            ws,
            {"m": "create_series", "p": [session_id, "sds_1", "s1", "sds_sym_1", "D", TV_CANDLE_COUNT]},
        )

    def on_message(ws: websocket.WebSocket, message: str) -> None:
        for item in parse_socket_message(message):
            if isinstance(item, dict) and item.get("m") == "timescale_update":
                try:
                    for candle in item["p"][1]["sds_1"]["s"]:
                        values = candle["v"]
                        candles.append(
                            {
                                "ts": values[0],
                                "open": round(values[1], 4),
                                "high": round(values[2], 4),
                                "low": round(values[3], 4),
                                "close": round(values[4], 4),
                            }
                        )
                except (KeyError, IndexError, TypeError) as exc:
                    print(f"  WARN {key}: parse error: {exc}")
                ws.close()
            elif isinstance(item, str) and item.startswith("~h~"):
                ws.send(f"~m~{len(item)}~m~{item}")

    def on_error(_ws: websocket.WebSocket, error: Exception) -> None:
        if "429" in str(error):
            got_429.set()
        else:
            print(f"  WARN {key}: {str(error)[:120]}")
        done.set()

    def on_close(_ws: websocket.WebSocket, *_args: Any) -> None:
        if candles:
            store_candles(key, candles)
        done.set()

    app = websocket.WebSocketApp(
        tradingview_url(tv_symbol),
        header=TV_HEADERS,
        on_open=on_open,
        on_message=on_message,
        on_error=on_error,
        on_close=on_close,
    )
    thread = threading.Thread(target=app.run_forever, kwargs={"ping_timeout": 15}, daemon=True)
    thread.start()
    done.wait(timeout=25)
    if got_429.is_set():
        if attempt <= MAX_RETRIES:
            wait = RETRY_DELAY * attempt
            print(f"  RETRY {key:10s}  429 rate-limited - retry {attempt}/{MAX_RETRIES} in {wait:.0f}s")
            time.sleep(wait)
            fetch_tradingview_symbol(key, tv_symbol, attempt + 1)
        else:
            print(f"  FAIL {key:10s}  gave up after {MAX_RETRIES} retries (still 429)")
    elif key not in state.market:
        print(f"  WARN {key:10s}  no data")


def fetch_all_symbols() -> None:
    """Fetch daily candles for every configured TradingView symbol."""
    print("Fetching live data from TradingView...\n")
    threads = []
    for key, symbol in SYMBOLS.items():
        thread = threading.Thread(target=fetch_tradingview_symbol, args=(key, symbol), daemon=True)
        thread.start()
        threads.append(thread)
        time.sleep(LAUNCH_DELAY)
    for thread in threads:
        thread.join(timeout=60)
    print(f"\n  Got data for {len(state.market)}/{len(SYMBOLS)} symbols\n")


def thin_border() -> Border:
    """Return a thin grey cell border."""
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_cell(
    cell: Any,
    value: Any,
    is_pct: bool = False,
    fill: PatternFill | None = None,
    number_format: str | None = None,
) -> None:
    """Apply value, font, border, alignment, fill, and number format to a cell."""
    cell.value = value
    cell.font = BODY_FONT
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    elif is_pct and isinstance(value, (int, float)):
        cell.number_format = "+0.00%;-0.00%;0.00%"
    elif isinstance(value, float):
        cell.number_format = "#,##0.0000"


def apply_index_cell(cell: Any, value: Any, is_pct: bool = False, fill: PatternFill | None = None) -> None:
    """Apply index-style number formatting to a cell."""
    fmt = "+0.00%;-0.00%;0.00%" if is_pct else "#,##0.00"
    number_format = fmt if isinstance(value, (int, float)) else None
    apply_cell(cell, value, is_pct=is_pct, fill=fill, number_format=number_format)


def style_title_cell(cell: Any, value: str | None = None) -> None:
    """Style a merged section title cell."""
    cell.value = value
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = thin_border()


def style_subheader_cell(cell: Any, label: str) -> None:
    """Style a column sub-header cell."""
    cell.value = label
    cell.font = SUBHDR_FONT if cell.column > 1 else BODY_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = thin_border()


def change_percent(close: float | None, open_value: float | None) -> float | None:
    """Calculate intraday percentage change from open to close."""
    if open_value and open_value != 0 and close is not None:
        return round((close - open_value) / open_value, 6)
    return None


def direction_fill(close: float | None, open_value: float | None) -> PatternFill | None:
    """Return green or red fill based on close versus open."""
    if open_value is None or close is None:
        return None
    return UP_FILL if close >= open_value else DOWN_FILL


def net_value_fill(value: float | None) -> PatternFill | None:
    """Return green or red fill based on sign of a net flow value."""
    if value is None:
        return None
    return UP_FILL if value >= 0 else DOWN_FILL


def parse_crore_value(raw: Any) -> float | None:
    """Parse NSDL or Moneycontrol crore values, including parenthesized negatives."""
    if raw is None:
        return None
    text = str(raw).strip().replace(",", "")
    if not text:
        return None
    if text.startswith("(") and text.endswith(")"):
        return -round(float(text[1:-1]), 2)
    return round(float(text), 2)


def parse_moneycontrol_date_key(key: str) -> str:
    """Convert Moneycontrol day keys such as 02Jun to ISO dates."""
    return datetime.strptime(f"{key}{datetime.now().year}", "%d%b%Y").strftime("%Y-%m-%d")


def http_get(url: str, as_json: bool = False, extra_headers: dict[str, str] | None = None) -> Any:
    """Perform an HTTP GET request and return text or JSON."""
    headers = {**HTTP_HEADERS, **(extra_headers or {})}
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=20) as response:
        body = response.read().decode("utf-8", errors="replace")
    return json.loads(body) if as_json else body


def parse_moneycontrol_js_var(html: str, var_name: str) -> dict | None:
    """Extract a JSON object assigned to a JavaScript variable in HTML."""
    match = re.search(rf"var\s+{var_name}\s*=\s*(\{{.*?\}});", html, re.S)
    if not match:
        return None
    return json.loads(match.group(1))


def latest_moneycontrol_key(data: dict[str, Any]) -> str:
    """Return the most recent day key from a Moneycontrol data dictionary."""
    return max(
        data.keys(),
        key=lambda key: datetime.strptime(f"{key}{datetime.now().year}", "%d%b%Y"),
    )


def fetch_moneycontrol_fii() -> dict[str, Any]:
    """Fetch cash and F&O FII nets from the Moneycontrol fiidii widget."""
    html = http_get(MC_FII_WIDGET)
    output: dict[str, Any] = {}
    net_block = parse_moneycontrol_js_var(html, "fiiDiiNetData")
    if net_block:
        day = latest_moneycontrol_key(net_block)
        output["date"] = parse_moneycontrol_date_key(day)
        output["fii_net"] = parse_crore_value(net_block[day]["fii"])
        output["dii_net"] = parse_crore_value(net_block[day]["dii"])
    fno_mapping = (
        ("fiiDiiIndFutData", "fii_fut_net"),
        ("fiiDiiIndOptData", "fii_opt_net"),
        ("fiiDiiStkFutData", "fii_stk_fut"),
        ("fiiDiiStkOptData", "fii_stk_opt"),
    )
    for source, target in fno_mapping:
        block = parse_moneycontrol_js_var(html, source)
        if block:
            day = latest_moneycontrol_key(block)
            output.setdefault("date", parse_moneycontrol_date_key(day))
            output[target] = parse_crore_value(block[day])
    if "fii_fut_net" in output and "fii_stk_fut" in output:
        output["fii_fut_net"] = round(output["fii_fut_net"] + output["fii_stk_fut"], 2)
    if "fii_opt_net" in output and "fii_stk_opt" in output:
        output["fii_opt_net"] = round(output["fii_opt_net"] + output["fii_stk_opt"], 2)
    return output


def fetch_nse_cash_fii() -> dict[str, Any]:
    """Fetch cash-segment FII and DII nets from NSE as a fallback source."""
    rows = http_get(
        NSE_FII_URL,
        as_json=True,
        extra_headers={"Referer": "https://www.nseindia.com/reports/fii-dii"},
    )
    output: dict[str, Any] = {}
    for row in rows:
        category = row.get("category", "")
        if category == "FII/FPI":
            output["fii_net"] = parse_crore_value(row["netValue"])
            output["date"] = datetime.strptime(row["date"], "%d-%b-%Y").strftime("%Y-%m-%d")
        elif category == "DII":
            output["dii_net"] = parse_crore_value(row["netValue"])
    return output


def fetch_nsdl_fpi_sebi() -> dict[str, Any]:
    """Fetch FPI equity and debt net investment from the NSDL daily report."""
    soup = BeautifulSoup(http_get(NSDL_FPI_URL), "lxml")
    report_date = None
    equity_net = None
    debt_total = 0.0
    debt_subtotals = 0
    for row in soup.find_all("tr"):
        cells = [cell.get_text(strip=True) for cell in row.find_all(["td", "th"])]
        if len(cells) < 2:
            continue
        if re.match(r"\d{2}-[A-Za-z]{3}-\d{4}", cells[0]):
            if report_date is not None:
                break
            report_date = datetime.strptime(cells[0], "%d-%b-%Y").strftime("%Y-%m-%d")
            continue
        if cells[0] == "Sub-total" and len(cells) >= 4:
            net = parse_crore_value(cells[3])
            if equity_net is None:
                equity_net = net
            elif debt_subtotals < FPI_DEBT_SUBTOTALS:
                debt_total += net
                debt_subtotals += 1
    return {
        "date": report_date,
        "equity": equity_net,
        "debt": round(debt_total, 2) if debt_subtotals else None,
    }


def fetch_fii_dii() -> None:
    """Populate FII/DII fields used by the USDINR sheet."""
    state.fii.clear()
    print("Fetching FII/DII data (Moneycontrol + NSDL)...\n")
    try:
        moneycontrol = fetch_moneycontrol_fii()
        state.fii.update(moneycontrol)
        print(f"  OK Moneycontrol  cash FII {moneycontrol.get('fii_net')}  DII {moneycontrol.get('dii_net')}")
        if moneycontrol.get("fii_fut_net") is not None:
            print(
                f"     F&O FII Fut {moneycontrol.get('fii_fut_net')}  "
                f"FII Opt {moneycontrol.get('fii_opt_net')}"
            )
    except OSError as exc:
        print(f"  WARN Moneycontrol FII widget failed: {exc}")
        try:
            nse = fetch_nse_cash_fii()
            state.fii.update(nse)
            print(f"  OK NSE fallback    FII {nse.get('fii_net')}  DII {nse.get('dii_net')}")
        except OSError as fallback_exc:
            print(f"  WARN NSE FII fallback failed: {fallback_exc}")
    try:
        sebi = fetch_nsdl_fpi_sebi()
        if sebi.get("date"):
            state.fii.setdefault("date", sebi["date"])
        state.fii["equity"] = sebi.get("equity")
        state.fii["debt"] = sebi.get("debt")
        print(f"  OK NSDL FPI SEBI  EQUITY {sebi.get('equity')}  DEBT {sebi.get('debt')}")
    except OSError as exc:
        print(f"  WARN NSDL FPI (EQUITY/DEBT) failed: {exc}")
    if state.market and state.fii.get("date"):
        tradingview_date = next(iter(state.market.values()))["date"]
        if state.fii["date"] != tradingview_date:
            print(f"  INFO FII report date {state.fii['date']} (TV candles: {tradingview_date})")
    print()


def excel_date_serial(date_str: str) -> int:
    """Convert YYYY-MM-DD to an Excel date serial number."""
    target = datetime.strptime(date_str, "%Y-%m-%d").date()
    epoch = date(1899, 12, 30)
    return (target - epoch).days


def next_append_row(worksheet: Worksheet) -> int:
    """Return the next empty row after the last populated row in column A."""
    last_row = 1
    for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=False):
        if row[0].value is not None:
            last_row = row[0].row
    return last_row + 1


def market_close_utc(date_str: str, market: str) -> datetime:
    """Return the UTC datetime when the given market's daily session closes.

    Combines the exchange-local calendar date with the UTC close hour so that
    column A of each sheet shows the exact moment the candle was sealed rather
    than the arbitrary script run time.
    """
    close_h = session_close_utc_hour(market)
    return datetime.strptime(date_str, "%Y-%m-%d").replace(hour=close_h)


def write_timestamp_cell(worksheet: Worksheet, row: int, close_dt: datetime) -> None:
    """Write market session close date and time (UTC) into column A."""
    cell = worksheet.cell(row, 1, close_dt)
    cell.number_format = "DD-MMM-YY HH:MM"
    cell.font = BODY_FONT
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_date_cell(worksheet: Worksheet, row: int, serial: int) -> None:
    """Write a legacy date-only value into column A."""
    cell = worksheet.cell(row, 1, serial)
    cell.number_format = "DD-MMM-YY"
    cell.font = BODY_FONT
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_open_close_pair(
    worksheet: Worksheet,
    row: int,
    start_col: int,
    data: dict[str, Any] | None,
) -> None:
    """Write open, close, and change percent for one instrument block."""
    fill = direction_fill(data["close"], data["open"]) if data else None
    apply_cell(worksheet.cell(row, start_col), data["open"] if data else None, fill=fill)
    apply_cell(worksheet.cell(row, start_col + 1), data["close"] if data else None, fill=fill)
    apply_cell(
        worksheet.cell(row, start_col + 2),
        change_percent(data["close"], data["open"]) if data else None,
        is_pct=True,
        fill=fill,
    )


def write_field(
    worksheet: Worksheet,
    row: int,
    column: int,
    key: str,
    field: str = "close",
    is_percent: bool = False,
    fill: PatternFill | None = None,
) -> None:
    """Write one market field from state into a worksheet cell."""
    record = state.market.get(key)
    value = record.get(field) if record else None
    apply_cell(worksheet.cell(row, column), value, is_pct=is_percent, fill=fill)


def append_dxy(workbook: Any, date_str: str) -> None:
    """Append DXY and major FX pairs to the DXY sheet."""
    worksheet = workbook["DXY"]
    next_row = next_append_row(worksheet)
    write_timestamp_cell(worksheet, next_row, market_close_utc(date_str, "FX"))
    column = 2
    for key in DXY_PAIRS:
        record = state.market.get(key)
        open_value = record["open"] if record else None
        close_value = record["close"] if record else None
        fill = direction_fill(close_value, open_value) if record and close_value and open_value else None
        apply_cell(worksheet.cell(next_row, column), open_value, fill=fill)
        apply_cell(worksheet.cell(next_row, column + 1), close_value, fill=fill)
        column += 2
    print(f"  OK DXY sheet: row {next_row} at {state.run_at:%Y-%m-%d %H:%M} (market {date_str})")


def append_usdinr_fii_block(worksheet: Worksheet, row: int) -> None:
    """Write FII/DII and FPI SEBI columns on the USDINR sheet."""
    for column, key in FII_COLUMN_KEYS:
        value = state.fii.get(key)
        apply_cell(
            worksheet.cell(row, column),
            value,
            fill=net_value_fill(value),
            number_format="#,##0.00" if isinstance(value, (int, float)) else None,
        )


def append_usdinr(workbook: Any, date_str: str) -> None:
    """Append FX, commodities, and FII data to the USDINR sheet."""
    worksheet = workbook["USDINR"]
    next_row = next_append_row(worksheet)
    write_timestamp_cell(worksheet, next_row, market_close_utc(date_str, "FX"))
    dxy = state.market.get("DXY")
    dxy_fill = direction_fill(dxy["close"], dxy["open"]) if dxy else None
    write_field(worksheet, next_row, 2, "DXY", "open", fill=dxy_fill)
    write_field(worksheet, next_row, 3, "DXY", "close", fill=dxy_fill)
    apply_cell(
        worksheet.cell(next_row, 4),
        change_percent(dxy["close"], dxy["open"]) if dxy else None,
        is_pct=True,
        fill=dxy_fill,
    )
    inr = state.market.get("USDINR")
    inr_fill = direction_fill(inr["close"], inr["open"]) if inr else None
    for column, field in ((5, "open"), (6, "high"), (7, "low"), (8, "close")):
        write_field(worksheet, next_row, column, "USDINR", field, fill=inr_fill)
    apply_cell(
        worksheet.cell(next_row, 9),
        change_percent(inr["close"], inr["open"]) if inr else None,
        is_pct=True,
        fill=inr_fill,
    )
    npr = state.market.get("USDNPR")
    npr_fill = direction_fill(npr["close"], npr["open"]) if npr else None
    write_field(worksheet, next_row, 10, "USDNPR", "open", fill=npr_fill)
    write_field(worksheet, next_row, 11, "USDNPR", "close", fill=npr_fill)
    apply_cell(
        worksheet.cell(next_row, 12),
        change_percent(npr["close"], npr["open"]) if npr else None,
        is_pct=True,
        fill=npr_fill,
    )
    append_usdinr_fii_block(worksheet, next_row)
    for key, start_col in (("OIL", 19), ("XAUUSD", 22), ("XAUFUT", 25), ("XAGUSD", 28)):
        record = state.market.get(key)
        fill = direction_fill(record["close"], record["open"]) if record else None
        write_field(worksheet, next_row, start_col, key, "open", fill=fill)
        write_field(worksheet, next_row, start_col + 1, key, "close", fill=fill)
        apply_cell(
            worksheet.cell(next_row, start_col + 2),
            change_percent(record["close"], record["open"]) if record else None,
            is_pct=True,
            fill=fill,
        )
    for column in range(31, 34):
        apply_cell(worksheet.cell(next_row, column), None)
    print(f"  OK USDINR sheet: row {next_row} at {state.run_at:%Y-%m-%d %H:%M} (market {date_str})")


def fill_from_row_change(worksheet: Worksheet, row: int, change_col: int) -> PatternFill | None:
    """Derive a direction fill from an existing change-percent cell."""
    change = worksheet.cell(row, change_col).value
    if isinstance(change, (int, float)):
        return UP_FILL if change >= 0 else DOWN_FILL
    return None


def normalize_timestamp_cell(worksheet: Worksheet, row: int) -> None:
    """Ensure column A uses the datetime display format."""
    value = worksheet.cell(row, 1).value
    if value is None:
        return
    if isinstance(value, (int, float)):
        epoch = datetime(1899, 12, 30)
        value = epoch + timedelta(days=int(value))
    elif isinstance(value, str):
        value = datetime.strptime(value[:10], "%Y-%m-%d")
    if isinstance(value, datetime):
        write_timestamp_cell(worksheet, row, value)


def style_us_indices_data_row(worksheet: Worksheet, row: int) -> None:
    """Apply formatting to one US_Indices data row."""
    normalize_timestamp_cell(worksheet, row)
    for block_start, change_col in ((2, 4), (5, 7)):
        fill = fill_from_row_change(worksheet, row, change_col)
        for column in range(block_start, block_start + 3):
            value = worksheet.cell(row, column).value
            apply_index_cell(
                worksheet.cell(row, column),
                value,
                is_pct=(column == change_col),
                fill=fill,
            )


def style_us_indices_sheet(worksheet: Worksheet) -> None:
    """Apply header and data styling to the US_Indices sheet."""
    merged = [str(area) for area in worksheet.merged_cells.ranges]
    if "B1:D1" not in merged:
        worksheet.merge_cells("B1:D1")
    if "E1:G1" not in merged:
        worksheet.merge_cells("E1:G1")
    style_title_cell(worksheet["A1"])
    style_title_cell(worksheet["B1"], "NASDAQ-100")
    style_title_cell(worksheet["E1"], "S&P 500")
    style_subheader_cell(worksheet.cell(2, 1), "Date")
    for column, label in (
        (2, "Open"), (3, "Close"), (4, "Change %"),
        (5, "Open"), (6, "Close"), (7, "Change %"),
    ):
        style_subheader_cell(worksheet.cell(2, column), label)
    for row in range(3, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value is None:
            continue
        style_us_indices_data_row(worksheet, row)


def ensure_us_indices_sheet(workbook: Any) -> Worksheet:
    """Create or return the US_Indices worksheet with headers applied."""
    if US_SHEET not in workbook.sheetnames:
        worksheet = workbook.create_sheet(US_SHEET)
    else:
        worksheet = workbook[US_SHEET]
    if not worksheet.cell(1, 2).value:
        worksheet.cell(1, 2, "NASDAQ-100")
        worksheet.cell(1, 5, "S&P 500")
    style_us_indices_sheet(worksheet)
    return worksheet


def row_date_string(cell_value: Any) -> str | None:
    """Convert a worksheet date cell to YYYY-MM-DD."""
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y-%m-%d")
    if isinstance(cell_value, (int, float)):
        epoch = date(1899, 12, 30)
        return (epoch + timedelta(days=int(cell_value))).strftime("%Y-%m-%d")
    if cell_value is not None:
        return str(cell_value)[:10]
    return None


def migrate_us_data_from_indian_stock(workbook: Any) -> None:
    """Move legacy NDX/SPX values from Indian_Stock into US_Indices once."""
    us_sheet = ensure_us_indices_sheet(workbook)
    if us_sheet.max_row > 2:
        return
    indian_sheet = workbook["Indian_Stock"]
    for row in range(3, indian_sheet.max_row + 1):
        if indian_sheet.cell(row, 18).value is None:
            continue
        row_date = row_date_string(indian_sheet.cell(row, 1).value)
        if row_date is None:
            continue
        target_row = us_sheet.max_row + 1
        write_date_cell(us_sheet, target_row, excel_date_serial(row_date))
        for source_col, dest_col in ((18, 2), (19, 3), (20, 4), (21, 5), (22, 6), (23, 7)):
            value = indian_sheet.cell(row, source_col).value
            if value is not None:
                apply_index_cell(
                    us_sheet.cell(target_row, dest_col),
                    value,
                    is_pct=(dest_col in (4, 7)),
                )


def cleanup_indian_stock_us_columns(workbook: Any) -> None:
    """Clear US index columns that were previously stored on Indian_Stock."""
    worksheet = workbook["Indian_Stock"]
    for merged in list(worksheet.merged_cells.ranges):
        if str(merged).startswith(("R1", "S1", "T1", "U1", "V1", "W1")):
            worksheet.unmerge_cells(str(merged))
    for row in range(1, worksheet.max_row + 1):
        for column in range(18, 24):
            worksheet.cell(row, column).value = None


def append_us_indices(workbook: Any, date_str: str) -> None:
    """Append NASDAQ-100 and S&P 500 data to the US_Indices sheet."""
    worksheet = ensure_us_indices_sheet(workbook)
    next_row = next_append_row(worksheet)
    write_timestamp_cell(worksheet, next_row, market_close_utc(date_str, "US"))
    write_open_close_pair(worksheet, next_row, 2, state.market.get("NDX"))
    write_open_close_pair(worksheet, next_row, 5, state.market.get("SPX"))
    print(f"  OK {US_SHEET} sheet: row {next_row} at {state.run_at:%Y-%m-%d %H:%M} (market {date_str})")


def append_indian_stock(workbook: Any, date_str: str) -> None:
    """Append NIFTY and SENSEX data to the Indian_Stock sheet."""
    worksheet = workbook["Indian_Stock"]
    next_row = next_append_row(worksheet)
    write_timestamp_cell(worksheet, next_row, market_close_utc(date_str, "INDIA"))
    nifty = state.market.get("NIFTY")
    nifty_fill = direction_fill(nifty["close"], nifty["open"]) if nifty else None
    apply_cell(worksheet.cell(next_row, 2), nifty["open"] if nifty else None, fill=nifty_fill)
    apply_cell(worksheet.cell(next_row, 3), nifty["close"] if nifty else None, fill=nifty_fill)
    apply_cell(
        worksheet.cell(next_row, 4),
        change_percent(nifty["close"], nifty["open"]) if nifty else None,
        is_pct=True,
        fill=nifty_fill,
    )
    apply_cell(worksheet.cell(next_row, 5), None)
    apply_cell(worksheet.cell(next_row, 6), None)
    sensex = state.market.get("SENSEX")
    sensex_fill = direction_fill(sensex["close"], sensex["open"]) if sensex else None
    apply_cell(worksheet.cell(next_row, 7), sensex["open"] if sensex else None, fill=sensex_fill)
    apply_cell(worksheet.cell(next_row, 8), sensex["close"] if sensex else None, fill=sensex_fill)
    apply_cell(
        worksheet.cell(next_row, 9),
        change_percent(sensex["close"], sensex["open"]) if sensex else None,
        is_pct=True,
        fill=sensex_fill,
    )
    apply_cell(worksheet.cell(next_row, 10), None)
    apply_cell(worksheet.cell(next_row, 11), None)
    print(f"  OK Indian_Stock sheet: row {next_row} at {state.run_at:%Y-%m-%d %H:%M} (market {date_str})")


def load_workbook_file() -> Any | None:
    """Load the Excel workbook or print an error when missing."""
    try:
        return load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"ERROR File not found: {EXCEL_FILE}")
        print("   Make sure Forex_Insights.xlsx is in the same folder as this script.")
        return None


def run() -> None:
    """Fetch data and append rows to all configured workbook sheets."""
    fetch_all_symbols()
    fetch_fii_dii()
    if not state.market:
        print("ERROR No data fetched. Check your internet connection.")
        return
    state.run_at = datetime.now()
    fx_date    = state.market.get("DXY",    {}).get("date", "")
    india_date = state.market.get("NIFTY",  {}).get("date", "")
    us_date    = state.market.get("NDX",    {}).get("date", "")
    print(f"Run time   : {state.run_at:%Y-%m-%d %H:%M:%S} local")
    print(f"Market date: FX/Oil/Gold={fx_date}  India={india_date}  US={us_date}\n")
    season = "summer" if _dst_active() else "winter"
    for market, info in SESSION_INFO_NPT.items():
        open_t  = info[f"open_{season}"]
        close_t = info[f"close_{season}"]
        print(f"  {market:6s}  open {open_t:8s}  close {close_t:8s} NPT  {info['note']}")
    print()
    print(f"Opening: {EXCEL_FILE}")
    workbook = load_workbook_file()
    if workbook is None:
        return
    migrate_us_data_from_indian_stock(workbook)
    cleanup_indian_stock_us_columns(workbook)
    print("\nAppending data to sheets (new row every run)...\n")
    append_dxy(workbook, fx_date)
    append_usdinr(workbook, fx_date)
    append_indian_stock(workbook, india_date)
    append_us_indices(workbook, us_date)
    style_us_indices_sheet(workbook[US_SHEET])
    workbook.save(EXCEL_FILE)
    upload_to_drive()          # ← add this
    send_slack_summary(state)  # ← add this
    print(f"\nSaved: {EXCEL_FILE}")
    print("   Each run appends a new row with date and time in column A.")


def main() -> None:
    """Entry point for module execution."""
    run()


if __name__ == "__main__":
    main()